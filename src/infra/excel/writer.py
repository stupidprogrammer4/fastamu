import asyncio
from collections.abc import Sequence
from concurrent.futures import ProcessPoolExecutor
from typing import Any

from openpyxl import load_workbook

from .row import ExcelRow


class ExcelWriter:
    """Async Excel writer.

    Workbook I/O is blocking (CPU + disk), so each job runs in a separate
    process via a ``ProcessPoolExecutor`` — a queue fronting worker process(es)
    whose result we ``await``. Writing always means filling a given ``.xlsx``
    template and saving the result to ``output``. ``close()`` on app shutdown.
    """

    def __init__(self, max_workers: int = 1) -> None:
        self._pool = ProcessPoolExecutor(max_workers=max_workers)


    def _write_rows_job(
        self,
        template: str,
        output: str,
        sheet: str | None,
        start_row: int,
        titles: list[str] | None,
        rows: list[list[Any]],
    ) -> str:
        wb = load_workbook(template)
        ws = wb[sheet] if sheet else wb.active
        if ws is None:
            raise ValueError("workbook has no active worksheet")

        r = start_row
        if titles is not None:
            for c, title in enumerate(titles, start=1):
                ws.cell(row=r, column=c, value=title)
            r += 1
        for values in rows:
            for c, value in enumerate(values, start=1):
                ws.cell(row=r, column=c, value=value)
            r += 1

        wb.save(output)
        return output


    def _write_cell_job(self, template: str, output: str, sheet: str | None, cell: str, value: Any) -> str:
        wb = load_workbook(template)
        ws = wb[sheet] if sheet else wb.active
        if ws is None:
            raise ValueError("workbook has no active worksheet")
        ws[cell] = value
        wb.save(output)
        return output

    async def write_rows(
        self,
        template: str,
        output: str,
        rows: Sequence[ExcelRow],
        *,
        start_row: int,
        sheet: str | None = None,
        with_titles: bool = False,
    ) -> str:
        """Write a batch of rows into ``template`` starting at ``start_row``
        (1-based), saving to ``output``. With ``with_titles`` the header row is
        written first (so the data then starts at ``start_row + 1``)."""
        titles = type(rows[0]).titles() if (with_titles and rows) else None
        payload = [row.cells() for row in rows]
        loop = asyncio.get_running_loop()
        result = await loop.run_in_executor(
            self._pool, self._write_rows_job, template, output, sheet, start_row, titles, payload
        )
        return result

    async def write_cell(
        self,
        template: str,
        output: str,
        cell: str,
        value: Any,
        *,
        sheet: str | None = None,
    ) -> str:
        """Write a single ``value`` into a specific ``cell`` (e.g. ``"B3"``)."""
        loop = asyncio.get_running_loop()
        result = await loop.run_in_executor(
            self._pool, self._write_cell_job, template, output, sheet, cell, value
        )
        return result

    def close(self) -> None:
        self._pool.shutdown()
