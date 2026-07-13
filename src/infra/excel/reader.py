import asyncio
from concurrent.futures import ProcessPoolExecutor
from typing import Any, TypeVar

from openpyxl import load_workbook

from .row import ExcelRow

TRow = TypeVar("TRow", bound=ExcelRow)



class ExcelReader:
    """Async Excel reader — the counterpart to `ExcelWriter`.

    Reads run in a separate process via a ``ProcessPoolExecutor`` (a queue
    fronting worker process(es)) and are ``await``-ed. ``read_rows`` maps each
    sheet row onto a typed `ExcelRow` by field order, so pydantic validates and
    coerces the cell types. ``close()`` on app shutdown.
    """

    def __init__(self, max_workers: int = 1) -> None:
        self._pool = ProcessPoolExecutor(max_workers=max_workers)

    def _read_rows_job(
        self,
        path: str,
        sheet: str | None,
        start_row: int,
        n_cols: int,
        max_rows: int | None,
    ) -> list[list[Any]]:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet] if sheet else wb.active
            if ws is None:
                raise ValueError("workbook has no active worksheet")

            out: list[list[Any]] = []
            for i, row in enumerate(ws.iter_rows(min_row=start_row, max_col=n_cols, values_only=True)):
                if max_rows is not None and i >= max_rows:
                    break
                if all(v is None for v in row):  # stop at the first blank row
                    break
                out.append(list(row))
            return out
        finally:
            wb.close()


    def _read_cell_job(self, path: str, sheet: str | None, cell: str) -> Any:
        wb = load_workbook(path, data_only=True)
        try:
            ws = wb[sheet] if sheet else wb.active
            if ws is None:
                raise ValueError("workbook has no active worksheet")
            return ws[cell].value
        finally:
            wb.close()

    async def read_rows(
        self,
        path: str,
        row_model: type[TRow],
        *,
        start_row: int,
        sheet: str | None = None,
        limit: int | None = None,
    ) -> list[TRow]:
        """Read rows from ``start_row`` (1-based) into ``row_model`` instances,
        mapping columns to fields by order. Stops at the first blank row or
        after ``limit`` rows."""
        names = list(row_model.model_fields)
        loop = asyncio.get_running_loop()
        raw = await loop.run_in_executor(
            self._pool, self._read_rows_job, path, sheet, start_row, len(names), limit
        )
        return [row_model(**dict(zip(names, values))) for values in raw]

    async def read_cell(self, path: str, cell: str, *, sheet: str | None = None) -> Any:
        """Read a single cell value (e.g. ``"B3"``)."""
        loop = asyncio.get_running_loop()
        result = await loop.run_in_executor(self._pool, self._read_cell_job, path, sheet, cell)
        return result

    def close(self) -> None:
        self._pool.shutdown()
